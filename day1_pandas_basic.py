import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

df = pd.read_excel("dirty_sales_data.xlsx")


df = df.rename(columns={
    'OrderID': 'Order ID',
    'Customer': 'Customer Name',
    'Product': 'Product',
    'Qty': 'Quantity',
    'Price': 'Unit Price',
    'Total': 'Amount',
    'Date': 'Order Date',
    'City': 'City',
    'Email': 'Email'
})

# 自动计算金额
df['Amount'] = df['Quantity'] * df['Unit Price']
df = df.dropna(subset=['Amount'])

# 清洗重复 + 填充空值
df = df.drop_duplicates(subset=['Order ID', 'Customer Name', 'Product', 'Amount'])
df['Customer Name'] = df['Customer Name'].fillna('Unknown Customer')
df['City'] = df['City'].fillna('Unknown City')
df['Email'] = df['Email'].fillna('No Email')

# 日期格式化 + 金额取整（超级专业！）
df['Order Date'] = pd.to_datetime(df['Order Date'], errors='coerce')
df['Amount'] = df['Amount'].astype(int)

df = df[df['Amount'] >= 100]
df['Is Big Order'] = df['Amount'].apply(lambda x: 'Yes' if x > 1000 else 'No')
df = df.sort_values(by='Amount', ascending=False)
df['Amount'] = df['Amount'].astype(int)
df = df.reset_index(drop=True)

# ======== 终极美化保存 ========
with pd.ExcelWriter("FINAL_PROFESSIONAL_Sales_Report_Nov2025.xlsx", engine='openpyxl') as writer:
    df.to_excel(writer, index=False, sheet_name='Sales Report')

    ws = writer.sheets['Sales Report']

    # 1. 标题行加粗 + 蓝色背景 + 居中
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # 2. 自动列宽（再也不用手动拉了）
    for i, column in enumerate(ws.columns, 1):
        max_length = 0
        column_letter = get_column_letter(i)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # 3. 金额列货币格式（可选更牛）
    # for row in ws['F2:F{}'.format(ws.max_row)]:
    #     row[0].number_format = '$#,##0'

print("day1完成")